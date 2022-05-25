
from typing import Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from name_standardizer import NameStandardizer, SegmentCut, SegmentPick

VERSION_SHEET='版本'
JOB_SHEET='作业列表'
DEPENDENCY_SHEET='依赖列表'
GROUP_RELATION_SHEET='作业分组规则'
CYCLE_DEPENDENT_SHEET='周期自依赖映射'
STANDARDIZE_RULE_SHEET='作业名标准化规则'
EXCLUDE_NAME_SHEET='非自动生成根节点'

VERSION='v1.0'

class ScheduleConfig:

    class Cycle:
        def __init__(
            self,
            key, #作业执行周期
            cycle_name, #调度自依赖周期
            self_dependent_type, #调度自依赖日
            desc #描述
        ) -> None:
            self.key = key
            self.cycle_name = cycle_name
            self.self_dependent_type = self_dependent_type
            self.desc = desc
            
    class Job:
        def __init__(
            self,
            group: str, #作业分组编码
            theme: str, #作业主题编码
            sub_theme: str, #作业子主题编码
            name: str, #作业名称
            desc: str, #作业描述
            cycle: Any, #作业执行周期(D/MB/ME)
            self_dependent: bool, #是否自依赖上一周期(Y/N)
            raw_script: str, #脚本内容(支持多行)
            fail_retry_times: int, #失败重试次数
            fail_retry_interval: int, #失败重试间隔(分钟)
            max_dependent_level: int #最低依赖级别(DAG图的组依赖级别)
        ) -> None:
            """
            group: str 作业分组编码
            theme: str 作业主题编码
            sub_theme: str 作业子主题编码
            name: str 作业名称
            desc: str 作业描述
            cycle: str 作业执行周期(D/MB/ME)
            self_dependent: bool 是否自依赖上一周期(Y/N)
            raw_script: str 脚本内容(支持多行)
            fail_retry_times: int 失败重试次数
            fail_retry_interval: int 失败重试间隔(分钟)
            max_dependent_level: int 最低依赖级别(DAG图的组依赖级别)
            """
            self.group = group
            self.theme = theme
            self.sub_theme = sub_theme
            self.name = name
            self.desc = desc
            self.cycle = cycle
            self.self_dependent = self_dependent
            self.raw_script = raw_script
            self.fail_retry_times = fail_retry_times
            self.fail_retry_interval = fail_retry_interval
            self.max_dependent_level = max_dependent_level
            self.full_name = '{}_{}'.format(group, name)
    
    class Dependency:
        def __init__(
            self,
            group: str, #作业分组编码
            name: str, #作业名称
            dependency_group: str, #作业分组编码
            dependency_name: str #作业名称
        ) -> None:
            self.group = group
            self.name = name
            self.dependency_group = dependency_group
            self.dependency_name = dependency_name
            self.full_name = '{}_{}'.format(group, name)
            self.full_dependency_name = '{}_{}'.format(dependency_group, dependency_name)
    
    class GroupRelation:
        def __init__(
            self,
            group: str, #作业及工作流分组编码
            enclosing_group: str, #上级工作流分组编码
            enclosing_name_template: str, #工作流分组名称
            enclosing_desc_template: str #工作流分组描述
        ) -> None:
            self.group = group
            self.enclosing_group = enclosing_group
            self.enclosing_name_template = enclosing_name_template.replace('${', '{')
            self.enclosing_desc_template = enclosing_desc_template.replace('${', '{')
        
        def get_name_by_job(self, job):
            return self.enclosing_name_template.format(
                job_name=job.name,
                theme=job.theme,
                cycle=job.cycle.key,
                cycle_desc=job.cycle.desc,
                sub_theme=job.sub_theme,
                job_desc=job.desc)
        
        def get_name_by_detail(self, name, theme, cycle_key, sub_theme):
            return self.enclosing_name_template.format(
                job_name=name,
                theme=theme,
                cycle=cycle_key,
                sub_theme=sub_theme)
        
        def get_desc_by_job(self, job):
            return self.enclosing_desc_template.format(
                job_name=job.name,
                theme=job.theme,
                cycle=job.cycle.key,
                cycle_desc=job.cycle.desc,
                sub_theme=job.sub_theme,
                job_desc=job.desc)

    def __init__(self,
        cycles: dict[str, Cycle],
        jobs: dict[str, Job],
        dependencies: list[Dependency],
        group_relations: dict[str, GroupRelation],
        standardizer: NameStandardizer,
        exclude_names: set[str]
    ) -> None:
        self.cycles = cycles
        self.jobs = jobs
        self.dependencies = dependencies
        self.group_relations = group_relations
        self.standardizer = standardizer
        self.exclude_names = exclude_names

    def from_xlsx(path: str):
        config_book = openpyxl.load_workbook(path, read_only=True)
        version_sheet = config_book[VERSION_SHEET]
        if version_sheet['A1'].value != VERSION:
            raise Exception('Version Error!')
        
        cycles: dict[str, ScheduleConfig.Cycle] = { }
        cycle_sheet: Worksheet = config_book[CYCLE_DEPENDENT_SHEET]
        for row in cycle_sheet.iter_rows(min_row=2, values_only=True):
            cycle = ScheduleConfig.Cycle(
                row[0], #作业执行周期
                row[1], #调度自依赖周期
                row[2], #调度自依赖日
                row[3], #描述
            )
            cycles[cycle.key] = cycle
        
        standardizer = NameStandardizer()
        standardize_rule_sheet: Worksheet = config_book[STANDARDIZE_RULE_SHEET]
        for row in standardize_rule_sheet.iter_rows(min_row=2, values_only=True):
            group = row[0]
            rule_type = row[1]
            if not rule_type:
                continue
            
            rule_regex = row[2]
            if rule_type in ('THEME', 'SUB_THEME'):
                is_match_partial = row[3].upper() == 'Y'
                pick_matched_begin = row[4]
                pick_matched_end = row[5]
            else:
                is_match_partial = None
                pick_matched_begin = None
                pick_matched_end = None

            if rule_type == 'CUT':
                rule_action = SegmentCut(rule_regex)
            elif rule_type == 'THEME':
                rule_action = SegmentPick(rule_regex, is_match_partial=is_match_partial, pick_start=pick_matched_begin, pick_end=pick_matched_end)
            elif rule_type == 'SUB_THEME':
                rule_action = SegmentPick(rule_regex, is_match_partial=is_match_partial, pick_start=pick_matched_begin, pick_end=pick_matched_end)
            
            standardizer.add_group_rule(group, rule_type, rule_action)
        
        jobs: dict[str, ScheduleConfig.Job] = { }
        job_sheet: Worksheet = config_book[JOB_SHEET]
        for row in job_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cut_result, theme, sub_theme = standardizer.standardize(row[0], row[1])
            job = ScheduleConfig.Job(
                row[0], #作业分组编码
                theme, #作业主题编码
                sub_theme, #作业子主题编码
                cut_result, #作业名称
                row[2], #作业描述
                cycles[row[3]], #作业执行周期(D/MB/ME)
                row[4].upper() == 'Y', #是否自依赖上一周期(Y/N)->bool
                row[5], #脚本内容(支持多行)
                row[6], #失败重试次数
                row[7], #失败重试间隔(分钟)
                row[8] #最低依赖级别(DAG图的组依赖级别)
            )
            jobs[job.full_name] = job
        
        dependencies: list[ScheduleConfig.Dependency] = []
        dependency_sheet: Worksheet = config_book[DEPENDENCY_SHEET]
        for row in dependency_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            dependency = ScheduleConfig.Dependency(
                row[0], #作业分组编码
                row[1], #作业名称
                row[2], #作业分组编码
                row[3], #作业名称
            )
            dependencies.append(dependency)
        
        group_relations: dict[str, ScheduleConfig.GroupRelation] = { }
        group_relations_sheet: Worksheet = config_book[GROUP_RELATION_SHEET]
        for row in group_relations_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            group_relation = ScheduleConfig.GroupRelation(
                row[0], #作业及工作流分组编码
                row[1], #上级工作流分组编码
                row[2], #工作流分组名称
                row[3], #工作流分组描述
            )
            group_relations[group_relation.group] = group_relation

        exclude_names_sheet: Worksheet = config_book[EXCLUDE_NAME_SHEET]
        exclude_names = set([
            row[0] for row in exclude_names_sheet.iter_rows(min_row=2, values_only=True)
        ])
        exclude_names.add('HEAD_M_MANUAL')
        exclude_names.add('HEAD_D_MANUAL')

        schedule_config = ScheduleConfig(cycles, jobs, dependencies, group_relations, standardizer, exclude_names)

        return schedule_config
    
    def get_group_relation(self, group, job):
        group_relation = self.group_relations.get(group)
        if group_relation:
            return [group_relation.group, group_relation.get_name_by_job(job), group_relation.get_desc_by_job(job)]
        return None

    def get_job_path(self, job: Job, group=None):
        if not group:
            group = job.group
        group_relation = self.group_relations.get(group)
        if group_relation:
            return self.get_job_path(job, group_relation.enclosing_group) + '/' + group_relation.get_name_by_job(job)
        return ''

    def get_job_name_path(self, group, name, theme, cycle_key, sub_theme):
        group_relation = self.group_relations.get(group)
        if group_relation:
            return self.get_job_name_path(
                    group_relation.enclosing_group, name, theme, cycle_key, sub_theme
                ) + '/' + group_relation.get_name_by_detail(
                    name, theme, cycle_key, sub_theme
                )
        return ''

if __name__ == '__main__':
    schedule_config = ScheduleConfig.from_xlsx('D:/Working/dolphinscheduler/调度模板.xlsx')

    print(schedule_config.get_job_path(schedule_config.jobs['ODB_S01_LNLNSLNS']))