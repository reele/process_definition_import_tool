
from typing import Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from name_standardizer import NameStandardizer, SegmentCut, SegmentPick
from simple_log import print_and_log

VERSION_SHEET='版本'
JOB_SHEET='作业列表'
DEPENDENCY_SHEET='依赖列表'
GROUP_DEPENDENT_LEVEL_SHEET='分组最高被依赖级别'
GROUP_RELATION_SHEET='作业分组规则'
CYCLE_DEPENDENT_SHEET='周期自依赖映射'
STANDARDIZE_RULE_SHEET='作业名标准化规则'
EXCLUDE_NAME_SHEET='非自动生成根节点'
MODIFIABLE_PROJECTS_SHEET='可操作项目'

# VERSION='v1.1' # 增加多项目依赖
# VERSION='v1.2' # 增加多项目分组
# VERSION='v1.3' # 增加调度分组多环境配置
VERSION='v1.4' # 增加自定义路径及项目保护

class ScheduleConfig:

    class Cycle:
        def __init__(
            self,
            key, #作业执行周期
            cycle_name, #调度自依赖周期
            self_dependent_type, #调度自依赖日
            desc #描述
        ) -> None:
            self.key = key
            self.cycle_name = cycle_name
            self.self_dependent_type = self_dependent_type
            self.desc = desc
            
    class Job:
        def __init__(
            self,
            project_name: str, #项目名称
            group: str, #作业分组编码
            theme: str, #作业主题编码
            sub_theme: str, #作业子主题编码
            name: str, #作业名称
            desc: str, #作业描述
            cycle: Any, #作业执行周期(D/MB/ME)
            self_dependent: bool, #是否自依赖上一周期(Y/N)
            raw_script: str, #脚本内容(支持多行)
            fail_retry_times: int, #失败重试次数
            fail_retry_interval: int, #失败重试间隔(分钟)
            max_dependent_level: int, #最低依赖级别(DAG图的组依赖级别)
            worker_group: int, #工作组
            environment: int, #环境
            enabled: str #是否启用
        ) -> None:
            """
            project_name: str 项目名称
            group: str 作业分组编码
            theme: str 作业主题编码
            sub_theme: str 作业子主题编码
            name: str 作业名称
            desc: str 作业描述
            cycle: str 作业执行周期(D/MB/ME)
            self_dependent: bool 是否自依赖上一周期(Y/N)
            raw_script: str 脚本内容(支持多行)
            fail_retry_times: int 失败重试次数
            fail_retry_interval: int 失败重试间隔(分钟)
            max_dependent_level: int 最低依赖级别(DAG图的组依赖级别)
            worker_group: 工作组
            environment: 环境
            enabled: str #启用类型
            """
            self.project_name = project_name
            self.group = group
            self.theme = theme
            self.sub_theme = sub_theme
            self.desc = desc
            self.cycle = cycle
            self.self_dependent = self_dependent
            self.raw_script = raw_script
            self.fail_retry_times = fail_retry_times
            self.fail_retry_interval = fail_retry_interval
            self.max_dependent_level = int(max_dependent_level)
            self.worker_group = worker_group
            self.environment = environment
            self.full_name = '{}_{}'.format(group, name)
            self.enabled = enabled
            if '/' in name:
                rindex = name.rindex('/')
                self.path = name[:rindex]
                self.name = name[rindex+1:]
            else:
                self.name = name
                self.path = None
        
        def is_full_path(self):
            return self.path is not None
    
    class Dependency:
        def __init__(
            self,
            project: str, #项目名称
            group: str, #作业分组编码
            name: str, #作业名称
            dependency_project: str, #项目名称
            dependency_group: str, #作业分组编码
            dependency_name: str, #作业名称
            highest_dependent_level: str #最高依赖级别
        ) -> None:
            self.project = project
            self.group = group
            self.name = name
            self.dependency_project = dependency_project
            self.dependency_group = dependency_group
            self.dependency_name = dependency_name
            self.full_name = '{}_{}'.format(group, name)
            self.full_dependency_name = '{}_{}'.format(dependency_group, dependency_name)
            if not highest_dependent_level:
                self.highest_dependent_level = 0
            else:
                self.highest_dependent_level = int(highest_dependent_level)
    
    class GroupRelation:
        def __init__(
            self,
            group: str, #作业及工作流分组编码
            enclosing_group: str, #上级工作流分组编码
            enclosing_name_template: str, #工作流分组名称
            enclosing_desc_template: str #工作流分组描述
        ) -> None:
            self.group = group
            self.enclosing_group = enclosing_group
            self.enclosing_name_template = enclosing_name_template.replace('${', '{')
            self.enclosing_desc_template = enclosing_desc_template.replace('${', '{')
        
        def get_name_by_job(self, job):
            return self.enclosing_name_template.format(
                job_name=job.name,
                theme=job.theme,
                cycle=job.cycle.key,
                cycle_desc=job.cycle.desc,
                sub_theme=job.sub_theme,
                job_desc=job.desc)
        
        def get_name_by_detail(self, name, theme, cycle_key, sub_theme):
            return self.enclosing_name_template.format(
                job_name=name,
                theme=theme,
                cycle=cycle_key,
                sub_theme=sub_theme)
        
        def get_desc_by_job(self, job):
            return self.enclosing_desc_template.format(
                job_name=job.name,
                theme=job.theme,
                cycle=job.cycle.key,
                cycle_desc=job.cycle.desc,
                sub_theme=job.sub_theme,
                job_desc=job.desc)

    def __init__(self,
        cycles: dict[str, Cycle],
        jobs: dict[str, Job],
        dependencies: list[Dependency],
        project_group_relations: dict[str, dict[str, GroupRelation]],
        standardizer: NameStandardizer,
        exclude_names: dict[str, set],
        group_highest_dependent_levels: dict[str, dict[str, set]],
        modifiable_projects: list[str]
    ) -> None:
        self.cycles = cycles
        self.jobs = jobs
        self.dependencies = dependencies
        self.project_group_relations = project_group_relations
        self.standardizer = standardizer
        self.exclude_names = exclude_names
        self.group_highest_dependent_levels = group_highest_dependent_levels
        self.modifiable_projects = modifiable_projects

    def from_xlsx(path: str):
        config_book = openpyxl.load_workbook(path, read_only=True)
        version_sheet = config_book[VERSION_SHEET]
        if version_sheet['A1'].value != VERSION:
            raise Exception('Version Error!')
        
        cycles: dict[str, ScheduleConfig.Cycle] = { }
        cycle_sheet: Worksheet = config_book[CYCLE_DEPENDENT_SHEET]
        for row in cycle_sheet.iter_rows(min_row=2, values_only=True):
            cycle = ScheduleConfig.Cycle(
                row[0], #作业执行周期
                row[1], #调度自依赖周期
                row[2], #调度自依赖日
                row[3], #描述
            )
            cycles[cycle.key] = cycle
        
        standardizer = NameStandardizer()
        standardize_rule_sheet: Worksheet = config_book[STANDARDIZE_RULE_SHEET]
        for row in standardize_rule_sheet.iter_rows(min_row=2, values_only=True):
            group = row[0]
            rule_type = row[1]
            if not rule_type:
                continue
            
            rule_regex = row[2]
            if rule_type in ('THEME', 'SUB_THEME'):
                is_match_partial = row[3].upper() == 'Y'
                pick_matched_begin = row[4]
                pick_matched_end = row[5]
            else:
                is_match_partial = None
                pick_matched_begin = None
                pick_matched_end = None

            if rule_type == 'CUT':
                rule_action = SegmentCut(rule_regex)
            elif rule_type == 'THEME':
                rule_action = SegmentPick(rule_regex, is_match_partial=is_match_partial, pick_start=pick_matched_begin, pick_end=pick_matched_end)
            elif rule_type == 'SUB_THEME':
                rule_action = SegmentPick(rule_regex, is_match_partial=is_match_partial, pick_start=pick_matched_begin, pick_end=pick_matched_end)
            
            standardizer.add_group_rule(group, rule_type, rule_action)
        
        jobs: dict[str, ScheduleConfig.Job] = { }
        job_sheet: Worksheet = config_book[JOB_SHEET]
        row_index = 2
        for row in job_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cut_result, theme, sub_theme = standardizer.standardize(row[1], row[2])

            enabled = None

            if row[12] == 'ON':
                enabled = 'YES'
            elif row[12] == 'OFF':
                enabled = 'NO'
            else:
                print_and_log("配置错误", "第[{}]行: 操作类型 [{}] 错误".format(row_index, row[12]))
                raise ValueError("配置错误, 已中止.")


            job = ScheduleConfig.Job(
                row[0], #作业所属项目
                row[1], #作业分组编码
                theme, #作业主题编码
                sub_theme, #作业子主题编码
                cut_result, #作业名称
                row[3], #作业描述
                cycles[row[4]], #作业执行周期(D/MB/ME)
                row[5].upper() == 'Y', #是否自依赖上一周期(Y/N)->bool
                row[6], #脚本内容(支持多行)
                row[7], #失败重试次数
                row[8], #失败重试间隔(分钟)
                row[9], #最低依赖级别(DAG图的组依赖级别)
                row[10], #工作组
                row[11], #环境
                enabled #启用标志
            )
            jobs[job.full_name] = job

            row_index += 1
        
        dependencies: list[ScheduleConfig.Dependency] = []
        dependency_sheet: Worksheet = config_book[DEPENDENCY_SHEET]
        for row in dependency_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            dependency = ScheduleConfig.Dependency(
                row[0], #项目名称
                row[1], #作业分组编码
                row[2], #作业名称
                row[3], #项目名称
                row[4], #作业分组编码
                row[5], #作业名称
                row[6], #最高依赖级别
            )
            dependencies.append(dependency)
        
        # group_relations: dict[str, ScheduleConfig.GroupRelation] = { }
        project_group_relations: dict[str, dict[str, ScheduleConfig.GroupRelation]] = { }
        group_relations_sheet: Worksheet = config_book[GROUP_RELATION_SHEET]
        for row in group_relations_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            project = row[0]
            group_relations = project_group_relations.get(project)
            if group_relations is None:
                group_relations = {}
                project_group_relations[project] = group_relations
            
            group_relation = ScheduleConfig.GroupRelation(
                row[1], #作业及工作流分组编码
                row[2], #上级工作流分组编码
                row[3], #工作流分组名称
                row[4], #工作流分组描述
            )
            group_relations[group_relation.group] = group_relation


        exclude_names_sheet: Worksheet = config_book[EXCLUDE_NAME_SHEET]
        exclude_names: dict[str, set] = {}
        for row in exclude_names_sheet.iter_rows(min_row=2, values_only=True):
            project_name = row[0]
            process_path = row[1]
            process_path_set = exclude_names.get(project_name)
            if process_path_set is None:
                process_path_set = set()
                exclude_names[project_name] = process_path_set
            
            process_path_set.add(process_path)

        exclude_names['[dw_main][1.1]'].add('/HEAD_M_MANUAL')
        exclude_names['[dw_main][1.1]'].add('/HEAD_D_MANUAL')
        exclude_names['[dw_main][1.1]'].add('/HEAD_D_DDBOA')
        exclude_names['[dw_main][1.1]'].add('/HEAD_D_ALERT')

        group_highest_dependent_levels_sheet: Worksheet = config_book[GROUP_DEPENDENT_LEVEL_SHEET]

        group_highest_dependent_levels: dict[str, dict[str, int]] = {}
        for row in group_highest_dependent_levels_sheet.iter_rows(min_row=2, values_only=True):
            project_name = row[0]
            group_code = row[1]
            max_level = int(row[2])
            levels = group_highest_dependent_levels.get(project_name)
            if levels is None:
                levels = {}
                group_highest_dependent_levels[project_name] = levels
            
            levels[group_code] = max_level
        
        modifiable_projects_sheet: Worksheet = config_book[MODIFIABLE_PROJECTS_SHEET]
        modifiable_projects: list[str] = []
        for row in modifiable_projects_sheet.iter_rows(min_row=2, values_only=True):
            project_name = row[0]
            if project_name:
                modifiable_projects.append(project_name)

        schedule_config = ScheduleConfig(
            cycles,
            jobs,
            dependencies,
            project_group_relations,
            standardizer,
            exclude_names,
            group_highest_dependent_levels,
            modifiable_projects)

        return schedule_config
    
    # def get_group_relation(self, group, job):
    #     group_relation = self.group_relations.get(group)
    #     if group_relation:
    #         return [group_relation.group, group_relation.get_name_by_job(job), group_relation.get_desc_by_job(job)]
    #     return None

    def measure_job_path(self, job: Job, group=None):
        if job.is_full_path():
            return f'/{job.project_name}/{job.path}'
        
        if not group:
            group = job.group
        group_relations = self.project_group_relations.get(job.project_name)
        if group_relations is None:
            print_and_log("配置错误", "作业[{}.{}]未定义项目的分组规则".format(job.project_name, job.name))
            raise ValueError("配置错误, 已中止.")
        group_relation = group_relations.get(group)
        if group_relation:
            return self.measure_job_path(job, group_relation.enclosing_group) + '/' + group_relation.get_name_by_job(job)
        return f'/{job.project_name}'

    def measure_job_name_path(self, project_name, group, name, theme, cycle_key, sub_theme):
        group_relations = self.project_group_relations.get(project_name)
        if group_relations is None:
            print_and_log("配置错误", "项目[{}]未定义分组规则".format(project_name))
            raise ValueError("配置错误, 已中止.")
        group_relation = group_relations.get(group)
        if group_relation:
            return self.measure_job_name_path(
                    project_name,
                    group_relation.enclosing_group, name, theme, cycle_key, sub_theme
                ) + '/' + group_relation.get_name_by_detail(
                    name, theme, cycle_key, sub_theme
                )
        return f'/{project_name}'

if __name__ == '__main__':
    schedule_config = ScheduleConfig.from_xlsx('调度模板v1.4.xlsx')

    print(schedule_config.measure_job_path(schedule_config.jobs['ODB_S01_FLNJT02']))
